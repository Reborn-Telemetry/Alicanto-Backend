from django.shortcuts import render, redirect
from openpyxl import Workbook
from bus_signals.query_utils import matriz_km_diario_flota, format_date, daily_bus_km, monthly_bus_km, monthly_fleet_km, recorrido_mensual_año
from bus_signals.models import Bus, BatteryHealth, ChargeStatus, CellsVoltage, EcuState, Odometer, FusiCode
from . models import Prueba, DisponibilidadFlota, MatrizEnergiaFlotaHistorico, DailyMatrizKmAutoReport
from datetime import date, timedelta, timezone
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
import io
from reportlab.platypus import Spacer
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from django.http import FileResponse, HttpResponse, JsonResponse
from reportlab.lib.pagesizes import letter, landscape
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image
import xlwt
from io import BytesIO
import psycopg2
from django.contrib import messages
from django.db.models import Max
from datetime import datetime
from django.utils import timezone
import requests
from django.views.decorators.csrf import csrf_exempt
from openai import ChatCompletion
from judini import CodeGPTPlus
from django.contrib.auth.decorators import login_required
import pytz
from django.db.models import Max, Case, When, Value
from django.db.models.functions import ExtractMonth, ExtractDay, ExtractYear
from django.db.models import Q, Count, Sum, Max, F, Avg
no_update_list = ['24', '87', '61','87', '137', '132', '134', '133', '130', '129', '128', '131', '136', '135' ]

#funcion reporte matriz seleccionando mes y año pagina historicos.
# optimizada version nueva mas rapida depende de los datos del modelo 
import io
from django.http import FileResponse
import xlwt
from datetime import datetime
from .models import Bus, DailyMatrizKmAutoReport
import json

def recorrido_informe_recorrido_mensual_año_xls(request):
    from bus_signals.query_utils import recorrido_mensual_mes_año

    if request.method == 'GET':
        context = {
            'bus': Bus.objects.all(),  # Ajustar según tu modelo
            'meses2': [
                {'mes': 'Enero', 'numero': 1}, {'mes': 'Febrero', 'numero': 2}, {'mes': 'Marzo', 'numero': 3},
                {'mes': 'Abril', 'numero': 4}, {'mes': 'Mayo', 'numero': 5}, {'mes': 'Junio', 'numero': 6},
                {'mes': 'Julio', 'numero': 7}, {'mes': 'Agosto', 'numero': 8}, {'mes': 'Septiembre', 'numero': 9},
                {'mes': 'Octubre', 'numero': 10}, {'mes': 'Noviembre', 'numero': 11}, {'mes': 'Diciembre', 'numero': 12},
            ],
        }
        return render(request, 'reports/historicos.html', context)

    if request.method == 'POST':
        # Obtener los parámetros mes y año desde el formulario
        mes = int(request.POST['mes'])
        año = int(request.POST['año'])
        
        # Obtener los datos del método
        data = recorrido_mensual_mes_año(mes, año)

        # Crear el archivo Excel en memoria
        current_datetime = datetime.now()
        filename = f"Recorrido_Buses_{mes}_{año}.xls"
        buf = io.BytesIO()
        workbook = xlwt.Workbook(encoding='utf-8')
        worksheet = workbook.add_sheet('Report')

        # Escribir los encabezados
        headers = ["Mes", "Bus", "Kilometraje inicial", "Kilometraje final", "Recorrido"]
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header)

        # Escribir los datos
        row_num = 1
        for row in data:
            bus = row[0]
            initial_km = row[1]
            final_km = row[2]
            recorrido = final_km - initial_km
            worksheet.write(row_num, 0, mes)
            worksheet.write(row_num, 1, bus)
            worksheet.write(row_num, 2, initial_km)
            worksheet.write(row_num, 3, final_km)
            worksheet.write(row_num, 4, recorrido)
            row_num += 1

        # Guardar el archivo en el buffer
        workbook.save(buf)
        buf.seek(0)

        # Crear la respuesta HTTP para descargar el archivo
        response = HttpResponse(buf, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response
    #---------------------------------------------------------------------------------

    

# esta funcion entrega la matriz de kilometraje de mes seleccionado de toda la flota
def historical_data(request):
    if request.method == 'GET':
        context = {
            'bus': Bus.bus.all().exclude(id__in=no_update_list),
            'meses2': [{'mes': 'Enero', 'numero': 1}, {'mes': 'Febrero', 'numero': 2}, {'mes': 'Marzo', 'numero': 3},
                       {'mes': 'Abril', 'numero': 4}, {'mes': 'Mayo', 'numero': 5}, {'mes': 'Junio', 'numero': 6},
                       {'mes': 'Julio', 'numero': 7}, {'mes': 'Agosto', 'numero': 8}, {'mes': 'Septiembre', 'numero': 9},
                       {'mes': 'Octubre', 'numero': 10}, {'mes': 'Noviembre', 'numero': 11}, {'mes': 'Diciembre', 'numero': 12}],
        }
        return render(request, 'reports/historicos.html', context)

    if request.method == 'POST':
        mes = int(request.POST['mes'])
        año = int(request.POST['año'])
        buses = Bus.bus.all().exclude(id__in=no_update_list)

        # Crear el archivo Excel
        current_datetime = datetime.now()
        formatted_datetime = current_datetime.strftime("%d-%m-%Y")
        filename = f'matriz_km_diario_flota_historico_mes_{mes}-{año}.xls'
        buf = io.BytesIO()
        workbook = xlwt.Workbook(encoding='utf-8')
        worksheet = workbook.add_sheet('Report')

        # Escribir encabezados
        headers = ["Mes", "Bus"] + [str(day) for day in range(1, 32)]
        for col_index, header in enumerate(headers):
            worksheet.write(0, col_index, header)

        row_index = 1  # Comenzar desde la segunda fila

        for bus in buses:
            # Obtener datos del modelo `DailyMatrizKmAutoReport` para el bus, mes y año seleccionado
            matriz_data = (
                DailyMatrizKmAutoReport.objects
                .filter(bus=bus, mes=mes, año=año)
                .values('dia', 'max_odometer')
            )

            # Crear una fila para cada bus
            row = [mes, bus.bus_name]

            # Inicializar una lista con 31 días en 0
            km_values = [0] * 31

            # Reemplazar los valores por los kilómetros registrados en `DailyMatrizKmAutoReport`
            for data in matriz_data:
                dia = data['dia'] - 1  # Ajustar el índice para el arreglo (día 1 es índice 0)
                km_values[dia] = data['max_odometer']

            # Agregar los valores a la fila
            row.extend(km_values)

            # Escribir la fila en la hoja de cálculo
            for col_index, cell_data in enumerate(row):
                worksheet.write(row_index, col_index, cell_data)
            row_index += 1

        # Guardar el archivo Excel en el buffer
        workbook.save(buf)
        buf.seek(0)

        # Retornar el archivo Excel como respuesta
        return FileResponse(buf, as_attachment=True, filename=filename)

# esta funcion entrega los kilometrajes diarios durante un mes seleccionado y año     
def historic_bus_report(request):
    if request.method == 'GET':
        # Lógica para manejar la solicitud GET
        return render(request, 'reports/historicos.html')

    if request.method == 'POST':
        # Obtener los datos enviados por el formulario
        mes = int(request.POST['mes'])
        año = int(request.POST['año'])
        bus_id = int(request.POST['bus'])

        # Obtener el bus específico
        try:
            bus = Bus.bus.get(id=bus_id)
        except Bus.DoesNotExist:
            return render(request, 'reports/historicos.html', {'error': 'Bus no encontrado'})

        # Inicializamos la tabla con el encabezado
        table_data = [
            ["Mes", "Bus", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", 
             "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27", 
             "28", "29", "30", "31"]
        ]

        # Consultar los kilómetros por día para el bus seleccionado en el mes y año dados
        km_data = (
            DailyMatrizKmAutoReport.objects
            .filter(bus=bus, mes=mes, año=año)
            .values('dia', 'max_odometer')
        )

        # Inicializar una lista de 31 días con 0 km
        km_values = [0] * 31

        # Rellenar los valores de km para los días disponibles
        for data in km_data:
            dia = data['dia'] - 1  # Ajustar el índice para la lista (día 1 es índice 0)
            if 0 <= dia < 31:
                km_values[dia] = data['max_odometer']

        # Agregar los datos del bus a la fila
        row = [mes, bus.bus_name] + km_values
        table_data.append(row)

        # Crear el archivo Excel
        filename = f'km_diario_bus_id_{bus_id}_mes_{mes}_año_{año}.xls'
        buf = io.BytesIO()
        workbook = xlwt.Workbook(encoding='utf-8')
        worksheet = workbook.add_sheet('Report')

        # Escribir encabezados
        for row_index, row_data in enumerate(table_data):
            for col_index, cell_data in enumerate(row_data):
                worksheet.write(row_index, col_index, cell_data)

        # Guardar el archivo Excel en el buffer
        workbook.save(buf)
        buf.seek(0)

        # Retornar el archivo Excel como respuesta
        return FileResponse(buf, as_attachment=True, filename=filename)
        

@login_required(login_url='login')
def dashboard_disponibilidad_flota(request):
   headers = { 'User-Agent': 'Alicanto/1.0', }
   api_url = 'https://reborn.assay.cl/api/v1/fs_elec'
   response = requests.get(api_url, headers=headers)
   data = response.json()
   list_fs_bus = data['data']
   cant_fs = len(data['data'])
   total_flota = Bus.bus.exclude(id__in=no_update_list)
   total_flota = total_flota.count()
   operacion = total_flota - cant_fs
   context = {
      'fs': list_fs_bus,
      'cant_fs': cant_fs,
      'total_flota': total_flota,
      'operacion': operacion,
   }
   return render(request, 'reports/disponibilidad_flota.html', context)


@csrf_exempt
def chatbot(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            print(f"Received message: {data}")
            user_message = data.get("message")

            # Preparar la solicitud a la API de CodeGPT
            print("Preparing request to CodeGPT API...")
            url = "https://api.codegpt.co/api/v1/chat/completions"
            headers = {
                "Authorization": "Bearer sk-dc8a8b1f-53bd-4eea-bc5d-26bee19cd783",
                "accept": "application/json",
                "content-type": "application/json",
            }
            payload = {
                "agentId": "1a289149-d981-4a10-be8d-c7c690e737f5",
                "messages": [{"role": "user", "content": user_message}],
                "format": "json",
                "stream": False
            }
            print(f"Payload: {payload}")

            response = requests.post(url, headers=headers, json=payload)
            print(f"Response status: {response.status_code}")
            print(f"Response data: {response.text}")

            if response.status_code != 200:
                return JsonResponse({"error": "Error from API"}, status=response.status_code)

            response_data = response.json()
            bot_response = response_data["choices"][0]["message"]["completion"]
            return JsonResponse({"response": bot_response})

        except Exception as e:
            print(f"Error: {str(e)}")
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request method"}, status=400)


@login_required(login_url='login')
def energy_report(request):
    current_datetime = datetime.now()
    mes = current_datetime.month
    año = current_datetime.year

    # Excluir ciertos buses
    bus_list = Bus.bus.exclude(id__in=no_update_list)

    # Configurar la zona horaria de Santiago
    santiago_tz = pytz.timezone('Chile/Continental')

    # Organizar los datos por bus y día
    lista_datos_organizados = {bus.id: {'bus': bus.bus_name, 'datos': {}} for bus in bus_list}

    # Obtener los datos desde MatrizEnergiaFlotaHistorico
    registros = MatrizEnergiaFlotaHistorico.objects.filter(
        bus__in=bus_list,
        mes=mes,
        año=año
    ).order_by('dia')

    # Organizar los datos por bus y día
    for registro in registros:
        bus_id = registro.bus_id
        fecha = datetime(año, mes, registro.dia, tzinfo=santiago_tz).strftime('%Y-%m-%d')
        if bus_id in lista_datos_organizados:
            lista_datos_organizados[bus_id]['datos'][fecha] = registro.energia

    # Generación del archivo Excel
    buf = io.BytesIO()
    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('Reporte')
    filename = f'energia_flota_{mes}-{año}.xls'

    # Encabezado de la primera fila
    primer_dia_mes = datetime(año, mes, 1, tzinfo=santiago_tz)
    ultimo_dia_mes = (primer_dia_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    dias_mes = [primer_dia_mes + timedelta(days=d) for d in range((ultimo_dia_mes - primer_dia_mes).days + 1)]

    worksheet.write(0, 0, "Bus")
    for col_num, dia in enumerate(dias_mes, start=1):
        worksheet.write(0, col_num, dia.strftime('%d'))

    # Datos de los buses
    for row_num, data in enumerate(lista_datos_organizados.values(), start=1):
        worksheet.write(row_num, 0, data['bus'])
        for fecha, energia_total in data['datos'].items():
            col_num = next((i + 1 for i, dia in enumerate(dias_mes) if fecha == dia.strftime('%Y-%m-%d')), None)
            if col_num:
                worksheet.write(row_num, col_num, energia_total)

    workbook.save(buf)
    buf.seek(0)

    # Retornar el archivo Excel como respuesta
    return FileResponse(buf, as_attachment=True, filename=filename)

@login_required(login_url='login')
def reporte_soh_flota(request):
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    table_data = [
        ["Bus", "Kilometraje", 'Capacidad Bateria', 'Desgaste Bateria x km']
    ]
    filename = f'estado baterias flota :{formatted_datetime}.xls'
    buf = io.BytesIO()
    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('Report')

    # Obtener los nombres de los buses y el máximo ID de BatteryHealth para cada uno
    nombres_buses_max_id = BatteryHealth.battery_health.values('bus__bus_name').annotate(max_id=Max('id'))

    # Obtener los registros asociados a los buses con el ID máximo, sin duplicados
    baterias = BatteryHealth.battery_health.filter(id__in=nombres_buses_max_id.values('max_id')).select_related('bus').order_by('bus__bus_name', 'bus__lts_odometer')

    for bateria in baterias:
        if bateria.battery_health_value < 100:
            desgaste = round(bateria.bus.lts_odometer / (100 - bateria.battery_health_value), 2)
        else:
            desgaste = 'N/A'  # O cualquier otro valor que prefieras

        row = [bateria.bus.bus_name, bateria.bus.lts_odometer, f"{bateria.battery_health_value}%", desgaste]
        table_data.append(row)

    for row_index, row_data in enumerate(table_data):
        for col_index, cell_data in enumerate(row_data):
            worksheet.write(row_index, col_index, cell_data)

    workbook.save(buf)
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename=filename)   


@login_required(login_url='login')
def last_soh_report(request):
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    table_data = [
        ["Bus","Capacidad Bateria", "fecha"]
    ]
    filename = f'estado baterias flota :{formatted_datetime}.xls'
    buf = io.BytesIO()
    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('Report')


@login_required(login_url='login')
def daily_bus_km_report_pdf(request, pk):
    bus = Bus.bus.get(id=pk)
    result = daily_bus_km(pk)

    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    filename = f'km_diario_bus_{bus.bus_name}_{formatted_datetime}.pdf'
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    elements = []

    table_data = [
        ["Dia", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    ]

    # Crear una lista de listas con los datos de result organizados por mes
    data_by_month = [[] for _ in range(12)]  # 12 meses
    for item in result:
        for idx, value in enumerate(item[1:], start=1):
            if value is not None:
                data_by_month[idx - 1].append(value)
            else:
                data_by_month[idx - 1].append('')

    # Llenar la tabla con los datos organizados
    for i in range(31):  # 31 días
        row = [f"Día {i+1}"]
        for month_data in data_by_month:
            if i < len(month_data):
                row.append(month_data[i])
            else:
                row.append('')
        table_data.append(row)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), 'grey'),
        ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), 'lightgrey'),
        ('LINEBELOW', (0, 1), (-1, -1), 1, 'black')
    ])

    table = Table(table_data)
    table.setStyle(style)
    elements.append(table)

    image_path = 'static/img/REM.png'
    image = Image(image_path, width=80, height=80)
    elements.insert(0, image)

    doc.build(elements)

    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=filename)


"""
def historical_energy_report(request):
    if request.method == 'GET':
        context = {
            'bus': Bus.bus.all().exclude(id__in=no_update_list),
            'meses2': [{'mes': 'Enero', 'numero': 1}, {'mes': 'Febrero', 'numero': 2}, {'mes': 'Marzo', 'numero': 3},
                       {'mes': 'Abril', 'numero': 4}, {'mes': 'Mayo', 'numero': 5}, {'mes': 'Junio', 'numero': 6},
                       {'mes': 'Julio', 'numero': 7}, {'mes': 'Agosto', 'numero': 8}, {'mes': 'Septiembre', 'numero': 9},
                       {'mes': 'Octubre', 'numero': 10}, {'mes': 'Noviembre', 'numero': 11}, {'mes': 'Diciembre', 'numero': 12}],
        }

        return render(request, 'reports/historicos.html', context)

    if request.method == 'POST':
        mes = int(request.POST['mes'])
        año = int(request.POST['año'])
        bus_list = Bus.bus.exclude(id__in=no_update_list)

        santiago_tz = pytz.timezone('Chile/Continental')
        lista_datos_organizados = {bus.id: {'bus': bus.bus_name, 'datos': {}} for bus in bus_list}

        charge_data = ChargeStatus.charge_status.filter(bus_id__in=bus_list.values_list('id', flat=True)).order_by('TimeStamp')

        for item in charge_data:
            bus_id = item.bus_id
            if bus_id in lista_datos_organizados:
                if item.charge_status_value == 1:
                    lista_datos_organizados[bus_id].setdefault('rango_actual', []).append(item)
                elif item.charge_status_value == 0 and 'rango_actual' in lista_datos_organizados[bus_id]:
                    rango_actual = lista_datos_organizados[bus_id].pop('rango_actual')
                    lista_datos_organizados[bus_id].setdefault('rangos', []).append(rango_actual)

        for bus_id, data in lista_datos_organizados.items():
            rangos = data.get('rangos', [])
            datos_tabla = []
            for i, rango in enumerate(rangos, 1):
                fecha_inicio = rango[0].TimeStamp
                fecha_termino = rango[-1].TimeStamp
                soc_inicial = rango[0].soc_level
                soc_final = rango[-1].soc_level
                carga = soc_final - soc_inicial
                diferencia = fecha_termino - fecha_inicio
                diferencia_en_horas = diferencia.total_seconds() / 3600
                datos_tabla.append({
                    'rango': i,
                    'fecha_inicio': fecha_inicio,
                    'fecha_termino': fecha_termino,
                    'tiempo': round(diferencia_en_horas, 2),
                    'soc_inicial': soc_inicial,
                    'soc_final': soc_final,
                    'carga': carga,
                    'energia': (carga * 140) / 100,
                    'bus': data['bus']
                })

            primer_dia_mes = datetime(año, mes, 1, tzinfo=santiago_tz)
            ultimo_dia_mes = (primer_dia_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            dias_mes = [primer_dia_mes + timedelta(days=d) for d in range((ultimo_dia_mes - primer_dia_mes).days + 1)]

            tabla_energia = {fecha.strftime('%Y-%m-%d'): 0 for fecha in dias_mes}
            for dato in datos_tabla:
                fecha = dato['fecha_inicio'].strftime('%Y-%m-%d')
                if fecha in tabla_energia:
                    tabla_energia[fecha] += (dato['carga'] * 140) / 100

            for fecha, energia_total in tabla_energia.items():
                data['datos'][fecha] = round(energia_total, 2)

        # Generación del archivo Excel
        buf = io.BytesIO()
        workbook = xlwt.Workbook(encoding='utf-8')
        worksheet = workbook.add_sheet('Reporte')
        filename = f'energia_flota_{mes}-{año}.xls'

        # Encabezado de la primera fila
        worksheet.write(0, 0, "Bus")
        for col_num, dia in enumerate(dias_mes, start=1):
            worksheet.write(0, col_num, dia.strftime('%d'))

        # Datos de los buses
        for row_num, data in enumerate(lista_datos_organizados.values(), start=1):
            worksheet.write(row_num, 0, data['bus'])
            for fecha, energia_total in data['datos'].items():
                col_num = next((i + 1 for i, dia in enumerate(dias_mes) if fecha == dia.strftime('%Y-%m-%d')), None)
                if col_num:
                    worksheet.write(row_num, col_num, energia_total)

        workbook.save(buf)
        buf.seek(0)
        return FileResponse(buf, as_attachment=True, filename=filename)"""

@login_required(login_url='login')
def historic_soh(request):
    if request.method == 'GET':
         return render(request, 'reports/historicos.html')
    
    if request.method == 'POST':
        dbname = 'alicanto-db-dev'
        user = 'postgres'
        password = 'postgres'
        host = 'alicanto-db-v1.cyydo36bjzsy.us-west-1.rds.amazonaws.com'
        port = '5432'
        bus = request.POST['bus']
        año = request.POST['año']
        buf = io.BytesIO()
        filename = f'soh anual id:{bus}-año:{año}.pdf' 
        doc = SimpleDocTemplate(buf, pagesize=letter)
        elements = []
        table_data = [
        ["Dia", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        ]
        connection = psycopg2.connect(dbname=dbname, user=user, password=password, host=host, port=port)
        cursor = connection.cursor()

        query = """ WITH all_days AS (
             SELECT generate_series(1, 31) AS dia
                )
                SELECT
                    all_days.dia,
                    MAX(CASE WHEN mes = 1 THEN max_soh END) AS enero,
                    MAX(CASE WHEN mes = 2 THEN max_soh END) AS febrero,
                    MAX(CASE WHEN mes = 3 THEN max_soh END) AS marzo,
                    MAX(CASE WHEN mes = 4 THEN max_soh END) AS abril,
                    MAX(CASE WHEN mes = 5 THEN max_soh END) AS mayo,
                    MAX(CASE WHEN mes = 6 THEN max_soh END) AS junio,
                    MAX(CASE WHEN mes = 7 THEN max_soh END) AS julio,
                    MAX(CASE WHEN mes = 8 THEN max_soh END) AS agosto,
                    MAX(CASE WHEN mes = 9 THEN max_soh END) AS septiembre,
                    MAX(CASE WHEN mes = 10 THEN max_soh END) AS octubre,
                    MAX(CASE WHEN mes = 11 THEN max_soh END) AS noviembre,
                    MAX(CASE WHEN mes = 12 THEN max_soh END) AS diciembre
                FROM all_days
                LEFT JOIN (
                    SELECT
                        EXTRACT(DAY FROM "TimeStamp") AS dia,
                        bus_id,
                        battery_health_value as max_soh,
                        EXTRACT(MONTH FROM "TimeStamp") AS mes,
                        EXTRACT(YEAR FROM "TimeStamp") AS año
                    FROM (
                        WITH ranked_data AS (
                            SELECT
                                bus_id,
                                battery_health_value,
                                "TimeStamp",
                                ROW_NUMBER() OVER (PARTITION BY bus_id, DATE_TRUNC('day', "TimeStamp") ORDER BY battery_health_value DESC) AS rnk
                            FROM
                                bus_signals_batteryhealth
                        )
                        SELECT
                            bus_id,
                            battery_health_value,
                            "TimeStamp"
                        FROM
                            ranked_data
                        WHERE
                            rnk = 1
                        ORDER BY
                            bus_id, "TimeStamp" DESC
                    ) A
                    WHERE bus_id = %s

                ) A ON all_days.dia = A.dia
                WHERE año = %s
                GROUP BY all_days.dia; 
                """
        cursor.execute(query, (bus, año))
        result = cursor.fetchall()

        data_by_month = [[] for _ in range(12)]  # 12 meses
    for item in result:
        for idx, value in enumerate(item[1:], start=1):
            if value is not None:
                data_by_month[idx - 1].append(value)
            else:
                data_by_month[idx - 1].append('')

    # Llenar la tabla con los datos organizados
    for i in range(31):  # 31 días
        row = [f"Día {i+1}"]
        for month_data in data_by_month:
            if i < len(month_data):
                row.append(month_data[i])
            else:
                row.append('')
        table_data.append(row)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), 'grey'),
        ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), 'lightgrey'),
        ('LINEBELOW', (0, 1), (-1, -1), 1, 'black')
    ])

    table = Table(table_data)
    table.setStyle(style)
    elements.append(table)

    image_path = 'static/img/REM.png'
    image = Image(image_path, width=80, height=80)
    elements.insert(0, image)

    doc.build(elements)

    buf.seek(0)
        

    return FileResponse(buf, as_attachment=True, filename=filename)


@login_required(login_url='login')
def last_value_cells_deltas(request):
    dbname = 'alicanto-db-dev'
    user = 'postgres'
    password = 'postgres'
    host = 'alicanto-db-v1.cyydo36bjzsy.us-west-1.rds.amazonaws.com'
    port = '5432'   
    buf = io.BytesIO()
    bus = Bus.bus.all()
    filename = 'last_value_cells_deltas_flota.pdf' 
    doc = SimpleDocTemplate(buf, pagesize=letter)
    elements = []
    table_data = [[
            "Bus", "Max Voltaje", "Min Voltaje", "Prom Voltaje" ,"Δ Min/Max"
        ]]
    connection = psycopg2.connect(dbname=dbname, user=user, password=password, host=host, port=port)
    cursor = connection.cursor()
    for i in bus:
        bus_id = i.id
        query = """SELECT *
                   FROM bus_signals_cellsvoltage
                   WHERE bus_id = %s
                   ORDER BY "TimeStamp" DESC
                   LIMIT 1;"""
        cursor.execute(query, (bus_id,))
        results = cursor.fetchall()
        
        if results:
            row = [
                i.bus_name, 
                results[0][2], 
                results[0][3], 
                results[0][4], 
                round(results[0][2] - results[0][3], 4)
            ]
            table_data.append(row)
        else:
            print(f"No se encontraron resultados para el bus {i.bus_name}.")

    # Estilo de la tabla
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    # Agregar título
    styles = getSampleStyleSheet()
    title = Paragraph("Reporte de Últimos Valores de Voltaje de Celdas Flota", styles['Title'])
    elements.append(title)
    
    # Agregar la tabla al PDF
    elements.append(table)
    
    # Construir el PDF
    doc.build(elements)
    
    # Configurar el buffer para que FileResponse pueda leerlo
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=filename)

@login_required(login_url='login')
def last_value_cells_deltas_excel(request):
    dbname = 'alicanto-db-dev'
    user = 'postgres'
    password = 'postgres'
    host = 'alicanto-db-v1.cyydo36bjzsy.us-west-1.rds.amazonaws.com'
    port = '5432'   
    bus = Bus.bus.all()
    
    # Crear un nuevo Workbook de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Last Value Cells Deltas"
    
    # Especificar el nombre del archivo
    filename = 'last_value_cells_deltas_flota.xlsx' 
    
    # Encabezado de la hoja
    headers = ["Bus", "Max Voltaje", "Min Voltaje", "Prom Voltaje", "Δ Min/Max"]
    ws.append(headers)
    
    # Conectar a la base de datos
    connection = psycopg2.connect(dbname=dbname, user=user, password=password, host=host, port=port)
    cursor = connection.cursor()
    
    for i in bus:
        bus_id = i.id
        query = """SELECT *
                   FROM bus_signals_cellsvoltage
                   WHERE bus_id = %s
                   ORDER BY "TimeStamp" DESC
                   LIMIT 1;"""
        cursor.execute(query, (bus_id,))
        results = cursor.fetchall()
        
        if results:
            row = [
                i.bus_name, 
                results[0][2], 
                results[0][3], 
                results[0][4], 
                round(results[0][2] - results[0][3], 4)
            ]
            ws.append(row)
        else:
            print(f"No se encontraron resultados para el bus. {i.bus_name}.")
    
    # Guardar el archivo en un buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    # Retornar el archivo Excel como un FileResponse
    return FileResponse(buf, as_attachment=True, filename=filename)
    
@login_required(login_url='login')
def bus_performance_report_excel(request, pk):
    montly_result = monthly_bus_km(pk)
    bus = Bus.bus.get(pk=pk)
    current_datetime = timezone.now()
    mes_actual = current_datetime.strftime('%m')
    charge_data = ChargeStatus.charge_status.filter(bus_id=pk).order_by('TimeStamp')
    rangos = []
    rango_actual = []
    months_dict = {
    1: 'Enero',
    2: 'Febrero',
    3: 'Marzo',
    4: 'Abril',
    5: 'Mayo',
    6: 'Junio',
    7: 'Julio',
    8: 'Agosto',
    9: 'Septiembre',
    10: 'Octubre',
    11: 'Noviembre',
    12: 'Diciembre',
    }
    result_data = []
# Obtener el mes actual
    current_month = datetime.now().month
# Iterar sobre el rango correcto para cada mes
    if montly_result and len(montly_result[0]) > 0:
    # Iterar sobre todos los meses
        for month in range(1, 13):
         index = (month - 1) * 2 + 1  # Calcular el índice correspondiente en los resultados
         if index < len(montly_result[0]):
            value1 = montly_result[0][index]
            value2 = montly_result[0][index + 1] if index + 1 < len(montly_result[0]) else None
            difference = value2 - value1 if value2 is not None else None

            # Si el mes es el actual y solo hay un valor disponible (value1), indicar que está en curso
            if month == current_month and value2 is None:
                result_data.append({
                    'month': months_dict[month],
                    'value1': value1,
                    'value2': 'En curso',
                    'difference': 'Calculando'
                })
            elif value1 is not None and value2 is not None:
                result_data.append({
                    'month': months_dict[month],
                    'value1': value1,
                    'value2': value2,
                    'difference': difference if difference is not None else 'Calculando'
                })
            else:
                result_data.append({
                    'month': months_dict[month],
                    'value1': 0,
                    'value2': 0,
                    'difference': 0
                })
        else:
            result_data.append({
                'month': months_dict[month],
                'value1': 0,
                'value2': 0,
                'difference': 0
            })
    else:
    # Si no hay resultados, agregar todos los meses con ceros
        for month in range(1, 13):
            result_data.append({
            'month': months_dict[month],
            'value1': 0,
            'value2': 0,
            'difference': 0
        })


    for item in charge_data:
        if item.charge_status_value == 1:
            rango_actual.append(item)
        elif item.charge_status_value == 0:
            if rango_actual:
                rangos.append(rango_actual.copy())
                rango_actual.clear()
        else:
            continue
# Agregar el último rango si no termina con Estado 0.0
    if rango_actual:
        rangos.append(rango_actual)

    santiago_tz = pytz.timezone('Chile/Continental')
# Preparar los datos para la tabla y calcular acumulados
    datos_tabla = []
    for i, rango in enumerate(rangos, 1):
        fecha_inicio = rango[0].TimeStamp.strftime("%Y-%m-%d %H:%M:%S")
        fecha_termino = rango[-1].TimeStamp.strftime("%Y-%m-%d %H:%M:%S")
        soc_inicial = rango[0].soc_level
        soc_final = rango[-1].soc_level
        carga = soc_final - soc_inicial  # Resta de soc_level

        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d %H:%M:%S')
        fecha_termino_dt = datetime.strptime(fecha_termino, '%Y-%m-%d %H:%M:%S')

        fecha_inicio_dt_santiago = fecha_inicio_dt.replace(tzinfo=pytz.utc).astimezone(santiago_tz)
        fecha_termino_dt_santiago = fecha_termino_dt.replace(tzinfo=pytz.utc).astimezone(santiago_tz)
# Calcular la diferencia de tiempo en horas
        diferencia = fecha_termino_dt_santiago - fecha_inicio_dt_santiago
        diferencia_en_horas = diferencia.total_seconds() / 3600
        datos_tabla.append({
            'rango': i,
            'fecha_inicio': fecha_inicio_dt_santiago.strftime("%Y-%m-%d %H:%M:%S"),
            'fecha_termino': fecha_termino_dt_santiago.strftime("%Y-%m-%d %H:%M:%S"),
            'tiempo': round(diferencia_en_horas, 2),
            'soc_inicial': soc_inicial,
            'soc_final': soc_final,
            'carga': carga,
            'energia': (carga * 140) / 100 if bus.bus_series == 'Queltehue' else (carga * 280) / 100,
        })
    acumulado_mensual = {str(month).zfill(2): 0 for month in range(1, 13)} 
    for i in datos_tabla:
        fecha_inicio = i['fecha_inicio']
# fecha_inicio is in the format "YYYY-MM-DD HH:MM:SS"
        try:
            month = fecha_inicio[5:7]  # Extract month as a string
            acumulado_mensual[month] += i['energia']
        except ValueError:
        # Handle invalid month format (e.g., log or skip entry)
            print(f"Invalid month format found in fecha_inicio: {fecha_inicio}")
            pass
    monthly_totals = []
    for month, energy in acumulado_mensual.items():
        monthly_totals.append({'month': month, 'energy': round(energy, 2)})
    acu = round(sum(i['energia'] for i in datos_tabla), 2)

# Comenzamos con rendimiento --------------------------------------------------------->
    monthly_totals_dict = {item['month']: item['energy'] for item in monthly_totals}
    month_name_to_number = {
    'Enero': '01',
    'Febrero': '02',
    'Marzo': '03',
    'Abril': '04',
    'Mayo': '05',
    'Junio': '06',
    'Julio': '07',
    'Agosto': '08',
    'Septiembre': '09',
    'Octubre': '10',
    'Noviembre': '11',
    'Diciembre': '12'
    }
    combined_data = []
    calc_rendimiento = lambda energy, diff: round(float(energy) / float(diff), 2) if diff and energy and isinstance(diff, (int, float)) and isinstance(energy, (int, float)) else None
    for item in result_data:
        month_name = item['month']
        month_number = month_name_to_number[month_name]
        energy = float(monthly_totals_dict.get(month_number, 0))
    
        difference = item['difference']
        if isinstance(difference, str) or difference == 0:
            difference = None  # O manejar de otra forma, dependiendo de la lógica de negocio

        combined_data.append({
            'month': month_name,
            'difference': difference,
            'energy': energy,
            'rendimiento': calc_rendimiento(energy, difference)
        })
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Rendimiento Bus"
    filename = "Rendimiento Bus f'{bus.bus_name}'.xlsx"
    headers = ["Mes", "Recorrido KM", "Energia KWH", "Rendimiento KWH/KM"]
    ws.append(headers)
    for i in combined_data:
        if combined_data:
            ws.append([i['month'], i['difference'], i['energy'], i['rendimiento']])
        else:
            print(f"No se encontraron resultados para el bus. {i.bus_name}.")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=filename)

@login_required(login_url='login')
def switch_report_xls(request, pk):
    selected_bus = pk
    bus = Bus.bus.get(id=selected_bus)
    switch_state = EcuState.ecu_state.filter(bus_id=selected_bus).order_by('-TimeStamp')
    ranges = []
    rango = 0
    start_time = None
    for record in switch_state:
        # Accede a los campos del modelo directamente
        if record.sleep_state == 0 and start_time is None:
            start_time = record.TimeStamp  # Inicia el rango con la primera aparición de 0
        elif record.sleep_state != 0 and start_time is not None:
            end_time = record.TimeStamp  # Finaliza el rango con el primer valor diferente de 0
            rango += 1
            # Calcula la diferencia entre las fechas
            periodo = start_time - end_time
            # Convierte el periodo a minutos
            minutos = periodo.total_seconds() / 60
            hrs = minutos / 60
            # Almacena las fechas formateadas y el periodo en minutos
            ranges.append({
                "rango": rango,
                "inicio": format_date(end_time),
                "final": format_date(start_time),
                "periodo": f"{hrs:.2f} hrs"  # Formateamos el resultado en hrs
            })
            start_time = None  # Resetea para el siguiente rango
    wb = Workbook()
    ws = wb.active
    filename = f"Reporte Estado EcuSleep Bus '{bus.bus_name}'.xlsx"
    ws.title = f"Reporte Estado Ecu Sleep Bus '{bus.bus_name}'"
    headers = ["Rango", "Inicio", "Final", "Periodo Hrs"]
    ws.append(headers)
    for i in ranges:
        if ranges:
            ws.append([i['rango'], i['inicio'], i['final'], i['periodo']])
        else:
            print(f"No se encontraron resultados para el bus.")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=filename)

    
    return print("Generando reporte de switches...")
    
def drive_report(request, pk):
    selected_bus = pk
    bus = Bus.bus.get(id=selected_bus)
    drive_state = EcuState.ecu_state.filter(bus_id=selected_bus).order_by('-TimeStamp')
    ranges = []
    rango = 0
    start_time = None
    for record in drive_state:
        # Accede a los campos del modelo directamente
        if record.sleep_state == 9 and start_time is None:
            start_time = record.TimeStamp  # Inicia el rango con la primera aparición de 0
        elif record.sleep_state != 9 and start_time is not None:
            end_time = record.TimeStamp  # Finaliza el rango con el primer valor diferente de 0
            rango += 1
            # Calcula la diferencia entre las fechas
            periodo = start_time - end_time
            # Convierte el periodo a minutos
            minutos = periodo.total_seconds() / 60
            hrs = minutos / 60
            # Almacena las fechas formateadas y el periodo en minutos
            ranges.append({
                "rango": rango,
                "inicio": format_date(end_time),
                "final": format_date(start_time),
                "periodo": f"{hrs:.2f} hrs"  # Formateamos el resultado en hrs
            })
            start_time = None  # Resetea para el siguiente rango
    wb = Workbook()
    ws = wb.active
    filename = f"Reporte Estado Drive Bus '{bus.bus_name}'.xlsx"
    ws.title = f"Reporte Estado Drive Bus '{bus.bus_name}'"
    headers = ["Rango", "Inicio", "Final", "Periodo Hrs"]
    ws.append(headers)
    for i in ranges:
        if ranges:
            ws.append([i['rango'], i['inicio'], i['final'], i['periodo']])
        else:
            print(f"No se encontraron resultados para el bus.")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=filename)
    

@login_required(login_url='login')
def recorrido_mensual_bus_pdf(request, pk):
    bus = Bus.bus.get(id=pk)
    result = monthly_bus_km(pk)
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    filename = f'reporte_recorrido_mensual bus {bus.bus_name}_{formatted_datetime}.pdf'
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    elements = []

    # Crear la tabla con los datos de los meses y el recorrido
    table_data = [['Mes', 'Kilometraje Inicial', 'Kilometraje Final', 'Recorrido']]
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    idx_mes = 0  # Índice para recorrer los meses
    for item in result:
        # Ignorar el primer valor de la tupla
        item = item[1:]
        for i in range(len(item)):
            if i % 2 == 0:  # Índices impares para kilómetros iniciales
                mes = meses[idx_mes]
                km_inicial = item[i]
                km_final = item[i + 1] if i + 1 < len(item) else None
                recorrido = km_final - km_inicial if km_final is not None else ''
                table_data.append([mes, km_inicial, km_final, recorrido])
                idx_mes += 1  # Avanzar al siguiente mes

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('LINEBELOW', (0, 1), (-1, -1), 1, colors.black)
    ])

    # Crear la tabla y aplicar el estilo
    table = Table(table_data)
    table.setStyle(style)
    elements.append(table)

    image_path = 'static/img/REM.png'
    image = Image(image_path, width=80, height=80)
    elements.insert(0, image)

    # Construir el documento y guardar en el buffer
    doc.build(elements)

    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=filename)


@login_required(login_url='login')
def bus_historic_fusi(request, pk):
    bus = Bus.bus.get(id=pk)
    fusi_bus = FusiCode.fusi.filter(bus_id=pk).values('fusi_code').annotate(total=Count('fusi_code'))
    fusi_bus_complete = FusiCode.fusi.filter(bus_id=pk)
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    filename = f'reporte historico fusi_{bus.bus_name}_{formatted_datetime}.pdf'
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    elements = []
    table_data_2 = [['Fecha', 'Codigo Fusi', 'Kilometraje Ocurrencia']]
    table_data = [
        ["Codigo Fusi", "Ocurrencias"]
    ]
    for item in fusi_bus:
        table_data.append([item['fusi_code'], item['total']])
    
    for item2 in fusi_bus_complete:
        table_data_2.append([item2.TimeStamp.strftime("%d-%m-%Y %H:%M"), int(item2.fusi_code), item2.failure_odometer ])

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), 'grey'),
        ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), 'lightgrey'),
        ('LINEBELOW', (0, 1), (-1, -1), 1, 'black')
    ])

    # Crear la tabla y aplicar el estilo
    table = Table(table_data)
    table2 = Table(table_data_2)
    table2.setStyle(style)
    table.setStyle(style)
    spacer = Spacer(0, 20)
    elements.append(table)
    elements.append(spacer)
    elements.append(table2)
    
    image_path = 'static/img/REM.png'
    image = Image(image_path, width=80, height=80)
    elements.insert(0, image)

    doc.build(elements)

    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=filename)


#------------------------- funciones ok ---------------------------#
@login_required(login_url='login')
def historical_energy_report(request):
    if request.method == 'GET':
        context = {
            'bus': Bus.bus.all().exclude(id__in=no_update_list),
            'meses2': [{'mes': 'Enero', 'numero': 1}, {'mes': 'Febrero', 'numero': 2}, {'mes': 'Marzo', 'numero': 3},
                       {'mes': 'Abril', 'numero': 4}, {'mes': 'Mayo', 'numero': 5}, {'mes': 'Junio', 'numero': 6},
                       {'mes': 'Julio', 'numero': 7}, {'mes': 'Agosto', 'numero': 8}, {'mes': 'Septiembre', 'numero': 9},
                       {'mes': 'Octubre', 'numero': 10}, {'mes': 'Noviembre', 'numero': 11}, {'mes': 'Diciembre', 'numero': 12}],
        }
        return render(request, 'reports/historicos.html', context)

    if request.method == 'POST':
        mes = int(request.POST['mes'])
        año = int(request.POST['año'])
        bus_list = Bus.bus.exclude(id__in=no_update_list)

        santiago_tz = pytz.timezone('Chile/Continental')
        lista_datos_organizados = {bus.id: {'bus': bus.bus_name, 'datos': {}} for bus in bus_list}

        # Obtener los datos desde MatrizEnergiaFlotaHistorico
        registros = MatrizEnergiaFlotaHistorico.objects.filter(
            bus__in=bus_list,
            mes=mes,
            año=año
        ).order_by('dia')

        # Organizar los datos por bus y día
        for registro in registros:
            bus_id = registro.bus_id
            fecha = datetime(año, mes, registro.dia, tzinfo=santiago_tz).strftime('%Y-%m-%d')
            if bus_id in lista_datos_organizados:
                lista_datos_organizados[bus_id]['datos'][fecha] = registro.energia

        # Generación del archivo Excel
        buf = io.BytesIO()
        workbook = xlwt.Workbook(encoding='utf-8')
        worksheet = workbook.add_sheet('Reporte')
        filename = f'energia_flota_{mes}-{año}.xls'

        # Encabezado de la primera fila
        primer_dia_mes = datetime(año, mes, 1, tzinfo=santiago_tz)
        ultimo_dia_mes = (primer_dia_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        dias_mes = [primer_dia_mes + timedelta(days=d) for d in range((ultimo_dia_mes - primer_dia_mes).days + 1)]

        worksheet.write(0, 0, "Bus")
        for col_num, dia in enumerate(dias_mes, start=1):
            worksheet.write(0, col_num, dia.strftime('%d'))

        # Datos de los buses
        for row_num, data in enumerate(lista_datos_organizados.values(), start=1):
            worksheet.write(row_num, 0, data['bus'])
            for fecha, energia_total in data['datos'].items():
                col_num = next((i + 1 for i, dia in enumerate(dias_mes) if fecha == dia.strftime('%Y-%m-%d')), None)
                if col_num:
                    worksheet.write(row_num, col_num, energia_total)

        workbook.save(buf)
        buf.seek(0)
        return FileResponse(buf, as_attachment=True, filename=filename)

@login_required
def software_version(request):
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    filename = f'reporte versiones software :{formatted_datetime}.xls'
    buf = io.BytesIO()

    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('Report')

    table_data = [[
        'Bus', 'Mark', 'Jarvis', 'Vision', 'Fecha de actualización'
    ]]
    buses = Bus.bus.all()

    for i in buses:
        formatted_datetime = i.lts_update.strftime("%d/%m/%Y %H:%M") if i.lts_update else "sin actualizacion"
        row = [
            i.bus_name, i.mark, i.jarvis, i.vision, formatted_datetime
               ]
        table_data.append(row)
    
    style = xlwt.easyxf('font: bold on; align: horiz center')
    for row_num, row_data in enumerate(table_data):
        for col_num, cell_value in enumerate(row_data):
            worksheet.write(row_num, col_num, cell_value, style)

    workbook.save(buf)
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename=filename)

@login_required(login_url='login')
def matriz_km_diario_flota(request):
    current_datetime = datetime.now()
    mes = current_datetime.month
    año = current_datetime.year
    bus_list = Bus.bus.exclude(id__in=no_update_list)

    # Cabecera de la tabla
    table_data = [
        ["Mes", "Bus", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13",
         "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26","27",
         "28", "29", "30", "31"]
    ]
    
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    filename = f'matriz_km_diario_flota_{formatted_datetime}.xls'
    buf = BytesIO()
    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('Report')

    for bus in bus_list:
        # Consulta para obtener el odómetro máximo por cada día del mes en curso y año
        daily_kms = (
            DailyMatrizKmAutoReport.objects.filter(bus=bus, mes=mes, año=año)
            .values('dia')
            .annotate(max_odometer=Max('max_odometer'))
            .order_by('dia')
        )

        # Inicializar la fila con el nombre del mes y del bus
        row = [bus.bus_name]
        
        # Crear una lista con los valores de los días del mes (por defecto 0)
        km_values = [0] * 31
        
        # Asignar el valor del odómetro máximo en el día correspondiente
        for day_data in daily_kms:
            dia = day_data['dia'] - 1  # Restar 1 para ajustar el índice de la lista
            km_values[dia] = day_data['max_odometer'] if day_data['max_odometer'] is not None else 0

        # Agregar los valores de los días a la fila
        row.extend(km_values)
        
        # Agregar la fila a los datos de la tabla
        table_data.append([mes] + row)

    # Escribir los datos en la hoja de cálculo
    for row_index, row_data in enumerate(table_data):
        for col_index, cell_data in enumerate(row_data):
            worksheet.write(row_index, col_index, cell_data)

    # Guardar el archivo Excel en el buffer
    workbook.save(buf)
    buf.seek(0)

    # Retornar el archivo Excel como respuesta
    return FileResponse(buf, as_attachment=True, filename=filename)

@login_required
def recorrido_mensual_flota_xls(request):
    now = datetime.now()
    year = now.year
    result = recorrido_mensual_año(year)  # Asegúrate de que esta función esté correctamente definida
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    filename = f'reporte_recorrido_mensual_flota_{formatted_datetime}.xls'
    
    # Crear un buffer para almacenar el archivo Excel
    buf = BytesIO()

    # Crear un libro de trabajo (workbook) y una hoja de trabajo (worksheet)
    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('Recorrido Mensual Flota')

    # Estilo para la cabecera
    header_style = xlwt.easyxf('font: bold 1')

    # Encabezados de la tabla
    headers = ['Bus', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

    # Escribir los encabezados en la primera fila
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header, header_style)

    # Escribir los datos en las filas siguientes
    for row_num, item in enumerate(result, start=1):
        # Escribir el nombre del bus en la primera columna
        worksheet.write(row_num, 0, item[0])

        # Iterar sobre los valores de cada mes, calculando el recorrido (final - inicial) y escribir en las columnas correspondientes
        for col_num in range(1, len(item[1:]), 2):
            km_inicial = item[col_num]  # Valor del odómetro inicial
            km_final = item[col_num + 1] if col_num + 1 < len(item) else None  # Valor del odómetro final

            if km_inicial is not None and km_final is not None:
                recorrido = km_final - km_inicial  # Restar el valor final menos el inicial
            else:
                recorrido = 'N/A'  # Si no hay datos, usar 'N/A'

            # Escribir el recorrido en la columna correspondiente
            worksheet.write(row_num, col_num // 2 + 1, recorrido)

    # Guardar el archivo Excel en el buffer
    workbook.save(buf)
    buf.seek(0)

    # Retornar el archivo Excel como respuesta
    return FileResponse(buf, as_attachment=True, filename=filename)
@login_required
def recorrido_mensual_flota(request):
    now = datetime.now()
    year = now.year
    result = recorrido_mensual_año(year)  # Asegúrate de que esta función esté correctamente definida
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    filename = f'reporte_recorrido_mensual_flota_{formatted_datetime}.pdf'
    buf = BytesIO()

    doc = SimpleDocTemplate(buf, pagesize=landscape(letter))  # Use landscape mode to fit all months
    elements = []
    
    # Table headers: Bus name and each month twice for initial and final odometer readings
    table_data = [
        ['Bus', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    ]
    
    # Iterate over the result data
    for item in result:
        row = [item[0]]  # Start with the bus name, which should be the first item in the list
        for i in range(1, len(item), 2):  # Iterate over the months (2 steps: initial and final values)
            km_inicial = item[i]
            km_final = item[i + 1] if i + 1 < len(item) else None
            recorrido = km_final - km_inicial if km_final is not None and km_inicial is not None else 'N/A'
            row.append(recorrido)  # Append the calculated difference

        table_data.append(row)  # Add the row to the table data
    
    # Create and style the table
    table = Table(table_data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('LINEBELOW', (0, 1), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)
    
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=filename)

@login_required
def monthly_bus_report(request):
    report_data = monthly_fleet_km()
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    filename = f'reporte km mensual :{formatted_datetime}.pdf'
    buf = io.BytesIO()

    doc = SimpleDocTemplate(buf, pagesize=landscape(letter))
    elements = []

    # Define las cabeceras de la tabla
    table_data = [[
        'Bus', 'Ene I', 'Ene F', 'Feb I', 'Feb F', 'Mar I', 'Mar F', 'Abr I', 'Abr F',
        'May I', 'May F', 'Jun I', 'Jun F', 'Jul I', 'Jul F', 'Ago I', 'Ago F',
        'Sep I', 'Sep F', 'Oct I', 'Oct F', 'Nov I', 'Nov F', 'Dic I', 'Dic F'
    ]]

    # Agregar los datos de la consulta
    for entry in report_data:
        row = []
        for value in entry:
            row.append(str(value) if value is not None else '0')
        table_data.append(row)

    # Estilo de la tabla
    cell_width = 17
    font_size = 8  # Aumentar el tamaño de fuente si es necesario
    style = TableStyle([
        # Estilo para las cabeceras (fondo gris claro y texto negro)
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),  # Color de fondo de las celdas
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), font_size),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)  # Líneas de la tabla
    ])

    table = Table(table_data, colWidths=[cell_width * 1.5] * len(table_data[0]))
    table.setStyle(style)
    elements.append(table)

    # Agregar una imagen opcional en el encabezado
    image_path = 'static/img/REM.png'
    image_width = 150
    image_height = 150
    image = Image(image_path, width=image_width, height=image_height)
    elements.insert(0, image)

    # Generar el PDF
    doc.build(elements)

    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=filename)

@login_required
def pdf_report(request):
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    filename = f'reporte dia :{formatted_datetime}.pdf'
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    elements = []

    table_data = [
        ["Bus", "Sniffer", "Ultimo SOC", "Ultimo Odometro", "Fecha de actualizacion"]
    ]

    bus_list = Bus.bus.all()
    for bus in bus_list:
        formatted_datetime = bus.lts_update.strftime("%d/%m/%Y %H:%M") if bus.lts_update else "sin actualizacion"
        row = [bus.bus_name, bus.sniffer, str(bus.lts_soc), str(bus.lts_odometer) + 'km', formatted_datetime]
        table_data.append(row)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), 'grey'),
        ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), 'lightgrey'),
    ])

    table = Table(table_data)
    table.setStyle(style)
    elements.append(table)

    image_path = 'static/img/REM.png'
    image = Image(image_path, width=80, height=80)
    elements.insert(0, image)

    doc.build(elements)

    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=filename)

@login_required
def xls_report(request):
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    filename = f'reporte dia :{formatted_datetime}.xls'
    buf = io.BytesIO()

    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('Report')

    table_data = [
        ["Bus", "Sniffer", "Carga", "Kiometraje", "Fecha de actualizacion"]
    ]

    bus_list = Bus.bus.all()
    for bus in bus_list:
        formatted_datetime = bus.lts_update.strftime("%d/%m/%Y %H:%M") if bus.lts_update else "sin actualizacion"
        row = [bus.bus_name, bus.sniffer, str(bus.lts_soc), str(bus.lts_odometer), formatted_datetime]
        table_data.append(row)

    style = xlwt.easyxf('font: bold on; align: horiz center')

    for row_num, row_data in enumerate(table_data):
        for col_num, cell_value in enumerate(row_data):
            worksheet.write(row_num, col_num, cell_value, style)

    workbook.save(buf)
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename=filename)

@login_required
def monthly_bus_report_xls(request):
    report_data = monthly_fleet_km()
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%d-%m-%Y")
    filename = f'reporte km mensual :{formatted_datetime}.xls'
    buf = io.BytesIO()

    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('Report')

    table_data = [
        ['Bus', 'Ene I', 'Ene F', 'Feb I', 'Feb F', 'Mar I', 'Mar F', 'Abr I', 'Abr F',
         'May I', 'May F', 'Jun I', 'Jun F', 'Jul I', 'Jul F', 'Ago I', 'Ago F',
         'Sep I', 'Sep F', 'Oct I', 'Oct F', 'Nov I', 'Nov F', 'Dic I', 'Dic F']
    ]

    for entry in report_data:
        row = [str(value) if value is not None else '0' for value in entry]
        table_data.append(row)

    style = xlwt.easyxf('font: bold on; align: horiz center')

    for row_num, row_data in enumerate(table_data):
        for col_num, cell_value in enumerate(row_data):
            worksheet.write(row_num, col_num, cell_value, style)

    workbook.save(buf)
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename=filename)